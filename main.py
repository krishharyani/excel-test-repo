from openpyxl import load_workbook

XLSX_PATH = "report.xlsx"
SHEET_NAME = "FR-Detail"   # change if needed

# Set these once based on your template
LABEL_COL = "A"            # Business Line column
NODEID_COL = "B"           # NodeID column
VALUE_COL = "M"            # Example: 2024 YTD Actual (change to your target)

def nbsp_level(x) -> int:
    if x is None:
        return 0
    s = str(x)
    i = 0
    while i < len(s) and ord(s[i]) == 160:  # NBSP
        i += 1
    return i

def clean_label(x) -> str:
    if x is None:
        return ""
    return str(x).replace("\u00a0", " ").strip()

def parse_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in {"n/a", "#n/a"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "")
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return None

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Stack tracks the current parent at each level: stack[level] = node_id
    stack = {}
    children = {}  # parent_node_id -> [child_node_ids]

    for r in range(2, ws.max_row + 1):  # start from row 2 if row 1 is header
        label_raw = ws[f"{LABEL_COL}{r}"].value
        node_id = ws[f"{NODEID_COL}{r}"].value
        val = ws[f"{VALUE_COL}{r}"].value

        lvl = nbsp_level(label_raw)
        label = clean_label(label_raw)

        if not label or not node_id:
            continue

        # Find parent: the most recent node at the highest level less than current
        parent_id = None
        if lvl > 0:
            parent_lvl = max((l for l in stack.keys() if l < lvl), default=None)
            if parent_lvl is not None:
                parent_id = stack[parent_lvl]

        if parent_id:
            children.setdefault(parent_id, []).append(node_id)

        # Update stack: this node is now the current node at this level
        stack[lvl] = node_id
        # Clear deeper levels (they're no longer valid parents)
        for deeper in list(stack.keys()):
            if deeper > lvl:
                del stack[deeper]

        print(f"row={r} lvl={lvl} node={node_id} parent={parent_id} label={label}")

    print("\n--- Parent -> Children ---")
    for parent, kids in children.items():
        print(f"{parent}: {kids}")

if __name__ == "__main__":
    main()